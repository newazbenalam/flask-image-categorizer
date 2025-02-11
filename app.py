from flask import Flask, render_template, request, redirect, url_for
import os
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.filters import AutoFilter
from datetime import datetime

app = Flask(__name__)

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, 'static/input')
CARTOON_DIR = os.path.join(BASE_DIR, 'static/cartoon_images')
PERSON_DIR = os.path.join(BASE_DIR, 'static/real_person')
UNCATEGORY_DIR = os.path.join(BASE_DIR, 'static/uncategory')
LOG_FILE = os.path.join(BASE_DIR, 'Image_Categorization_Log.xlsx')

# Ensure category folders exist
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(CARTOON_DIR, exist_ok=True)
os.makedirs(PERSON_DIR, exist_ok=True)
os.makedirs(UNCATEGORY_DIR, exist_ok=True)

# Initialize Excel log file if not exists
def initialize_log():
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(["Image ID", "Humour", "Sarcastic", "Offensive", "Motivational", "Overall", "Category", "Confidence", "Languages", "Timestamp"])
        wb.save(LOG_FILE)
        wb.close()

# Get all images in input directory
def get_images():
    return [f for f in os.listdir(INPUT_DIR) if f.lower().endswith(('png', 'jpg', 'jpeg', 'gif'))]

# Log categorization action
def log_action(image_id, humour, sarcastic, offensive, motivational, overall, category, confidence, languages):
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        headers = ["Image ID", "Humour", "Sarcastic", "Offensive", "Motivational", "Overall", "Category", "Confidence", "Languages", "Timestamp"]
        ws.append(headers)
        ws.auto_filter.ref = "A1:J1"
        wb.save(LOG_FILE)
    else:
        wb = load_workbook(LOG_FILE)
        ws = wb.active
        if not ws.auto_filter.ref:
            ws.auto_filter.ref = "A1:J1"

    row = [
        image_id,
        humour,
        sarcastic,
        offensive,
        motivational,
        overall,
        category,
        confidence,
        languages,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ]
    ws.append(row)
    wb.save(LOG_FILE)


@app.route('/', methods=['GET', 'POST'])
def index():
    images = get_images()
    current_image = images[0] if images else None

    if request.method == 'POST':
        image_id = request.form.get("image_id")
        humour = request.form.get("humour")
        sarcastic = request.form.get("sarcastic")
        offensive = request.form.get("offensive")
        motivational = request.form.get("motivational")
        overall = request.form.get("overall")
        category = request.form.get("category")
        confidence = request.form.get("confidence")
        
        # Capture multiple selected languages
        languages = request.form.getlist("languages")
        languages_str = ", ".join(languages)

        src_path = os.path.join(INPUT_DIR, image_id)
        if category == "cartoon":
            dest_dir = CARTOON_DIR
        elif category == "person":
            dest_dir = PERSON_DIR
        elif category == "uncategory":
            dest_dir = UNCATEGORY_DIR
        else:
            return redirect(url_for('index'))

        dest_path = os.path.join(dest_dir, image_id)
        shutil.move(src_path, dest_path)

        # Log action with selected languages
        log_action(image_id.split('.')[0], humour, sarcastic, offensive, motivational, overall, category, confidence, languages_str)

        return redirect(url_for('index'))

    return render_template('index.html', image=current_image)


if __name__ == '__main__':
    initialize_log()
    app.run(debug=True)
